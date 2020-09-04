# -*- coding: utf-8 -*-
"""
Created on Tue Jul 23 21:54:22 2019

@author: William
"""
data=[]
import openpyxl as xl
#straight data
wb = xl.load_workbook('data.xlsx')
ws = wb.get_sheet_by_name('Sheet1')
length0 = ws.max_row 
for i in range (1,length0+1):
    data.append([ws.cell(column=j,row=i).value for j in range (1,ws.max_column+1)])

wb = xl.load_workbook('data1.xlsx')
ws = wb.get_sheet_by_name('Sheet1')
length1 = ws.max_row
for i in range (1,length1+1):
    data.append([ws.cell(column=j,row=i).value for j in range (1,ws.max_column+1)])

#wb = xl.load_workbook('data2.xlsx')
#ws = wb.get_sheet_by_name('Sheet1')
#length2 = ws.max_row
#for i in range (1,length2+1):
#    data.append([ws.cell(column=j,row=i).value for j in range (1,ws.max_column+1)])
# 
#wb = xl.load_workbook('data3.xlsx')
#ws = wb.get_sheet_by_name('Sheet1')
#length3 = ws.max_row
#for i in range (1,length3+1):
#    data.append([ws.cell(column=j,row=i).value for j in range (1,ws.max_column+1)])
#
#wb = xl.load_workbook('data4.xlsx')
#ws = wb.get_sheet_by_name('Sheet1')
#length4 = ws.max_row
#for i in range (1,length4+1):
#    data.append([ws.cell(column=j,row=i).value for j in range (1,ws.max_column+1)])
#
#wb = xl.load_workbook('data5.xlsx')
#ws = wb.get_sheet_by_name('Sheet1')
#length5 = ws.max_row
#for i in range (1,length5+1):
#    data.append([ws.cell(column=j,row=i).value for j in range (1,ws.max_column+1)])
#
#wb = xl.load_workbook('data6.xlsx')
#ws = wb.get_sheet_by_name('Sheet1')
#length6 = ws.max_row
#for i in range (1,length6+1):
#    data.append([ws.cell(column=j,row=i).value for j in range (1,ws.max_column+1)])
#    
#wb = xl.load_workbook('data7.xlsx')
#ws = wb.get_sheet_by_name('Sheet1')
#length7 = ws.max_row
#for i in range (1,length7+1):
#    data.append([ws.cell(column=j,row=i).value for j in range (1,ws.max_column+1)])
#
#wb = xl.load_workbook('data8.xlsx')
#ws = wb.get_sheet_by_name('Sheet1')
#length8 = ws.max_row
#for i in range (1,length8+1):
#    data.append([ws.cell(column=j,row=i).value for j in range (1,ws.max_column+1)])
#
#wb = xl.load_workbook('data9.xlsx')
#ws = wb.get_sheet_by_name('Sheet1')
#length9 = ws.max_row
#for i in range (1,length9+1):
#    data.append([ws.cell(column=j,row=i).value for j in range (1,ws.max_column+1)])
#
#wb = xl.load_workbook('data10.xlsx')
#ws = wb.get_sheet_by_name('Sheet1')
#length10 = ws.max_row
#for i in range (1,length10+1):
#    data.append([ws.cell(column=j,row=i).value for j in range (1,ws.max_column+1)])
#%%
import numpy as np
l=[]
c=[]
r=[]
s=[]
t=[]
data_size = length0 + length1
# + length2 +length3 + length4 + length5 + length6 + length7 +length8 + length9 + length10
for i in range (0,data_size):
    l.append(data[i][0])
    c.append(data[i][1])
    r.append(data[i][2])
    s.append(data[i][3])
#    t.append(data[i][4])
L=np.array(l)[:,np.newaxis]
C=np.array(c)[:,np.newaxis]
R=np.array(r)[:,np.newaxis]
S=np.array(s)[:,np.newaxis]
#T=np.array(t)[:,np.newaxis]
X=np.hstack((L,C,R))
Y=np.hstack(S)
##%%
##preprocessing
##remove extreme data
#from sklearn.preprocessing import StandardScaler
#dele=[]
#for i in range(0,data_size):
#    if x[i][0]==0 or x[i][1]==0 or x[i][2]==0 : dele.append(i)
#    elif x[i][0]>200 or x[i][1]>200 or x[i][2]>200 : dele.append(i)
#    elif i>1 and x[i][1]>x[i][0]+15 or x[i][1]>x[i][2]+15:
#        if x[i][1]-x[i-1][1]<0.5 and x[i][1]-x[i-1][1]>-0.5:
#            y[i][0]=8
#            y[i-1][0]=8
#y=np.delete(y,dele,axis=0)
#x=np.delete(x,dele,axis=0)
##standradize
#scaler=StandardScaler()
#scaler.fit(x)
#print("mean:",scaler.mean_)
#print("scale:",scaler.scale_)
#%% svm
from sklearn.model_selection import train_test_split
from sklearn import svm
from sklearn.metrics import accuracy_score
from sklearn.model_selection import GridSearchCV
#from sklearn.preprocessing import scale
#x = scale(x)
x_train,x_test,y_train,y_test = train_test_split(X,Y,test_size=0.1,random_state=0)

param_grid = {'C': [100,500,1000,5000,7500,10000,20000],
              'gamma':[0.01,0.05,0.1,0.5,0.75,1,1.5, 2,5], }
gclf = GridSearchCV(svm.SVC(kernel='linear'), param_grid)
gclf.fit(x_train, y_train.ravel())
best_C=gclf.best_estimator_.C
best_gamma=gclf.best_estimator_.gamma
print("C_S:",best_C)
print("gamma_S:",best_gamma)
svc = svm.SVC(gamma=best_gamma, C=best_C)
svc.fit(x_train,y_train.ravel())

y_pdt_T = svc.predict(x_test)
acc_T=accuracy_score(y_pdt_T,y_test.ravel())
import pickle
filename="F74052081.sav"
pickle.dump(svc, open(filename, 'wb'))
pickle.dump(X, open("svm.pickle",'wb'))
# load model
l_model=pickle.load(open(filename,'rb'))
yp_l=l_model.predict(x_test)
print("acc load: %f " % accuracy_score(yp_l, y_test))

z = lambda x,y: (-svc.intercept_[0]-svc.coef_[0][0]*x-svc.coef_[0][1]*y) / svc.coef_[0][2]

tmp = np.linspace(-2,2,51)
x,y = np.meshgrid(tmp,tmp)
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
# Plot stuff.
fig = plt.figure()
ax  = fig.add_subplot(111, projection='3d')
ax.plot_surface(x, y, z(x,y))
ax.plot3D(X[Y==0,0], X[Y==0,1], X[Y==0,2],'ob')
ax.plot3D(X[Y==1,0], X[Y==1,1], X[Y==1,2],'sr')
plt.show()
#x_train,x_test,y_train,y_test = train_test_split(x,T,test_size=0.1,random_state=0)
#
#gclf = GridSearchCV(svm.SVC(kernel='rbf'), param_grid)
#gclf.fit(x_train, y_train.ravel())
#best_C=gclf.best_estimator_.C
#best_gamma=gclf.best_estimator_.gamma
#print("C_T:",best_C)
#print("gamma_T:",best_gamma)
#clf = svm.SVC(gamma=best_gamma, C=best_C)
#clf.fit(x_train,y_train.ravel())
#y_pdt_T = clf.predict(x_test)
#acc_T=accuracy_score(y_pdt_T,y_test.ravel())

##%% extratree
#from sklearn.model_selection import train_test_split
#from sklearn.ensemble import ExtraTreesClassifier
#from sklearn.metrics import accuracy_score
#
#x_train,x_test,y_train,y_test = train_test_split(x,T,test_size=0.1,random_state=0)
#clf = ExtraTreesClassifier(n_estimators=1,max_features=3)
#clf.fit(x_train,y_train.ravel())
#y_pdt_T = clf.predict(x)
#acc_T=accuracy_score(y_pdt_T,T.ravel())
#pdt_T=np.array(y_pdt_T)[:,np.newaxis]
#x_new=np.hstack((L,C,R,pdt_T))
#x_train,x_test,y_train,y_test = train_test_split(x_new,S,test_size=0.1,random_state=0)
#clf.fit(x_train, y_train.ravel())
#y_predict = clf.predict(x_test)
#acc=accuracy_score(y_predict,y_test.ravel())

##%% randomforest
#from sklearn.model_selection import train_test_split
#from sklearn.ensemble import RandomForestClassifier
#from sklearn.metrics import accuracy_score
#x_train, x_test, y_train, y_test = train_test_split(x, T, test_size = 0.1, random_state = 0)
#clf = RandomForestClassifier(n_estimators=2,max_features=3)
#clf.fit(x_train, y_train.ravel())
#y_pdt_T = clf.predict(x)
#acc=accuracy_score(y_pdt_T,T.ravel())
#acc_T=accuracy_score(y_pdt_T,T.ravel())
#pdt_T=np.array(y_pdt_T)[:,np.newaxis]
#x_new=np.hstack((L,C,R,pdt_T))
#x_train,x_test,y_train,y_test = train_test_split(x_new,S,test_size=0.1,random_state=0)
#clf.fit(x_train, y_train.ravel())
#y_predict = clf.predict(x_test)
#acc=accuracy_score(y_predict,y_test.ravel())

##%% decisiontree
#from sklearn.model_selection import train_test_split
#from sklearn import tree
#from sklearn.metrics import accuracy_score
#x_train, x_test, y_train, y_test = train_test_split(x, T, test_size = 0.1, random_state = 0)
#clf = tree.DecisionTreeClassifier()
#clf.fit(x_train, y_train.ravel())
#y_pdt_T = clf.predict(x)
#acc_T=accuracy_score(y_pdt_T,T.ravel())
#pdt_T=np.array(y_pdt_T)[:,np.newaxis]
#x_new=np.hstack((L,C,R,pdt_T))
#x_train,x_test,y_train,y_test = train_test_split(x_new,S,test_size=0.1,random_state=0)
#clf.fit(x_train, y_train.ravel())
#y_predict = clf.predict(x_test)
#acc=accuracy_score(y_predict,y_test.ravel())
##%% ada
#from sklearn.model_selection import train_test_split
#from sklearn import tree
#from sklearn.ensemble import AdaBoostClassifier
#from sklearn.metrics import accuracy_score
#x_train, x_test, y_train, y_test = train_test_split(x, T, test_size = 0.1, random_state = 0)
#clf = AdaBoostClassifier(tree.DecisionTreeClassifier())
#clf.fit(x_train, y_train.ravel())
#y_pdt_T = clf.predict(x)
#acc=accuracy_score(y_pdt_T,T.ravel())
#acc_T=accuracy_score(y_pdt_T,T.ravel())
#pdt_T=np.array(y_pdt_T)[:,np.newaxis]
#x_new=np.hstack((L,C,R,pdt_T))
#x_train,x_test,y_train,y_test = train_test_split(x_new,S,test_size=0.1,random_state=0)
#clf.fit(x_train, y_train.ravel())
#y_predict = clf.predict(x_test)
#acc=accuracy_score(y_predict,y_test.ravel())
##%% porter
#from sklearn_porter import Porter
#porter=Porter(clf,language='C')
#output=porter.export(embed_data=True)
#output_file_name="svm_moduleT.txt"
#with open(output_file_name,'w') as file_obj:
#    file_obj.write(output)