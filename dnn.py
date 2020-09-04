# -*- coding: utf-8 -*-
"""
Created on Sat Feb 29 17:05:51 2020

@author: William
"""
data=[]
import openpyxl as xl
#straight data
wb = xl.load_workbook('data.xlsx')
ws = wb.get_sheet_by_name('Sheet1')
length0 = ws.max_row 
for i in range (1,length0+1):
    data.append([ws.cell(column=j,row=i).value for j in range (1,ws.max_column+1)])

wb = xl.load_workbook('data1.xlsx')
ws = wb.get_sheet_by_name('Sheet1')
length1 = ws.max_row
for i in range (1,length1+1):
    data.append([ws.cell(column=j,row=i).value for j in range (1,ws.max_column+1)])

wb = xl.load_workbook('data2.xlsx')
ws = wb.get_sheet_by_name('Sheet1')
length2 = ws.max_row
for i in range (1,length2+1):
    data.append([ws.cell(column=j,row=i).value for j in range (1,ws.max_column+1)])
 
wb = xl.load_workbook('data3.xlsx')
ws = wb.get_sheet_by_name('Sheet1')
length3 = ws.max_row
for i in range (1,length3+1):
    data.append([ws.cell(column=j,row=i).value for j in range (1,ws.max_column+1)])

wb = xl.load_workbook('data4.xlsx')
ws = wb.get_sheet_by_name('Sheet1')
length4 = ws.max_row
for i in range (1,length4+1):
    data.append([ws.cell(column=j,row=i).value for j in range (1,ws.max_column+1)])

wb = xl.load_workbook('data5.xlsx')
ws = wb.get_sheet_by_name('Sheet1')
length5 = ws.max_row
for i in range (1,length5+1):
    data.append([ws.cell(column=j,row=i).value for j in range (1,ws.max_column+1)])

wb = xl.load_workbook('data6.xlsx')
ws = wb.get_sheet_by_name('Sheet1')
length6 = ws.max_row
for i in range (1,length6+1):
    data.append([ws.cell(column=j,row=i).value for j in range (1,ws.max_column+1)])
#%%
from sklearn.model_selection import train_test_split
from keras.utils import np_utils
import numpy as np
import random
l=[]
c=[]
r=[]
s=[]
t=[]
data_size = length0 + length1+length2 +length3 + length4 + length5 + length6# + length7 +length8 + length9 + length10
for i in range (0,data_size):
    l.append(data[i][0])
    c.append(data[i][1])
    r.append(data[i][2])
    s.append(data[i][3])
#    t.append(data[i][4])
for i in range (0,10):
    l.append(data[35][0]+random.random()-random.random())
    c.append(data[35][1]+random.random()-random.random())
    r.append(data[35][2]+random.random()-random.random())
    s.append(data[35][3])
for i in range (0,10):
    l.append(data[52][0]+random.random()-random.random())
    c.append(data[52][1]+random.random()-random.random())
    r.append(data[52][2]+random.random()-random.random())
    s.append(data[52][3])
for i in range (0,10):
    l.append(data[72][0]+random.random()-random.random())
    c.append(data[72][1]+random.random()-random.random())
    r.append(data[72][2]+random.random()-random.random())
    s.append(data[72][3])
for i in range (0,10):
    l.append(data[89][0]+random.random()-random.random())
    c.append(data[89][1]+random.random()-random.random())
    r.append(data[89][2]+random.random()-random.random())
    s.append(data[89][3])
for i in range (0,10):
    l.append(data[107][0]+random.random()-random.random())
    c.append(data[107][1]+random.random()-random.random())
    r.append(data[107][2]+random.random()-random.random())
    s.append(data[107][3])
for i in range (0,10):
    l.append(data[124][0]+random.random()-random.random())
    c.append(data[124][1]+random.random()-random.random())
    r.append(data[124][2]+random.random()-random.random())
    s.append(data[124][3])
L=np.array(l)[:,np.newaxis]
C=np.array(c)[:,np.newaxis]
#R=np.array(r)[:,np.newaxis]
S=np.array(s)[:,np.newaxis]
#T=np.array(t)[:,np.newaxis]
X=np.hstack((L,C))#,R))
Y=np.hstack(S)
x_train,x_test,y_train,y_test = train_test_split(X,Y,test_size=0.1,random_state=0)
y_train_onehot = np_utils.to_categorical(y_train)
y_test_onehot = np_utils.to_categorical(y_test)
#%% build model
from keras.models import Sequential
from keras.layers import Dense
model = Sequential()
model.add(Dense(units=128,
                input_dim=2,
                kernel_initializer='normal',
                activation='relu'))
model.add(Dense(units=128,
                activation='relu'))
model.add(Dense(units=128,
                activation='relu'))
model.add(Dense(units=128,
                activation='relu'))
model.add(Dense(units=128,
                activation='relu'))
model.add(Dense(units=128,
                activation='relu'))
model.add(Dense(units=64,
                activation='relu'))
model.add(Dense(units=32,
                activation='relu'))
model.add(Dense(units=16,
                activation='relu'))
model.add(Dense(units=8,
                activation='relu'))
model.add(Dense(units=7,
                kernel_initializer='normal',
                activation='softmax'))
print(model.summary())
#%% compile
import numpy
model.compile(loss='categorical_crossentropy',
              optimizer='adam',
              metrics=['accuracy'])
train_history = model.fit(x=x_train,
                          validation_split=0.1,
                          epochs=500,batch_size=20,
                          verbose=2,y=y_train_onehot)
scores = model.evaluate(x_test,y_test_onehot)
print('acc=',scores[1])
trans = numpy.argmax(y_train_onehot,axis=-1)
print(train_history.history.keys())
#%% plot
import matplotlib.pyplot as plt
#plt.plot(train_history.history['acc'])
#plt.plot(train_history.history['val_acc'])
#plt.title('model accuracy')
#plt.ylabel('accuracy')
#plt.xlabel('epoch')
#plt.legend(['train', 'test'], loc='upper left')
#plt.show()
def show_train_history(train_acc, test_acc):
    plt.plot(train_history.history[train_acc])
    plt.plot(train_history.history[test_acc])
    plt.title('Train history')
    plt.ylabel('Accuracy')
    plt.xlabel('Epoch')
    plt.legend(['train','test'],loc='upper left')
    plt.show()
show_train_history('accuracy','val_accuracy')
#%%
import pickle
filename="dnn.sav"
pickle.dump(model, open(filename, 'wb'))
#pickle.dump(X, open("svm.pickle",'wb'))