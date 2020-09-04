# -*- coding: utf-8 -*-
"""
Created on Tue Jul 23 21:54:22 2019

@author: William
"""
data=[]
import openpyxl as xl
#straight data
length=256
wb = xl.load_workbook('data/straight1.xlsx')
ws = wb.get_sheet_by_name('工作表1')
for i in range (2,length+1):
    data.append([ws.cell(column=j,row=i).value for j in range (1,ws.max_column+1)])

length=149
wb = xl.load_workbook('data/straight2.xlsx')
ws = wb.get_sheet_by_name('工作表1')
for i in range (2,length+1):
    data.append([ws.cell(column=j,row=i).value for j in range (1,ws.max_column+1)])

length=231
wb = xl.load_workbook('data/straight3.xlsx')
ws = wb.get_sheet_by_name('工作表1')
for i in range (2,length+1):
    data.append([ws.cell(column=j,row=i).value for j in range (1,ws.max_column+1)])
 
length=111
wb = xl.load_workbook('data/straight4.xlsx')
ws = wb.get_sheet_by_name('工作表1')
for i in range (2,length+1):
    data.append([ws.cell(column=j,row=i).value for j in range (1,ws.max_column+1)])

length=185
wb = xl.load_workbook('data/straight5.xlsx')
ws = wb.get_sheet_by_name('工作表1')
for i in range (2,length+1):
    data.append([ws.cell(column=j,row=i).value for j in range (1,ws.max_column+1)])

length=21
wb = xl.load_workbook('data/straight6.xlsx')
ws = wb.get_sheet_by_name('工作表1')
for i in range (2,length+1):
    data.append([ws.cell(column=j,row=i).value for j in range (1,ws.max_column+1)])

#right data
length=183
wb = xl.load_workbook('data/right1.xlsx')
ws = wb.get_sheet_by_name('工作表1')
for i in range (2,length+1):
    data.append([ws.cell(column=j,row=i).value for j in range (1,ws.max_column+1)])

length=186
wb = xl.load_workbook('data/right2.xlsx')
ws = wb.get_sheet_by_name('工作表1')
for i in range (2,length+1):
    data.append([ws.cell(column=j,row=i).value for j in range (1,ws.max_column+1)])

#left data
length=151
wb = xl.load_workbook('data/left1.xlsx')
ws = wb.get_sheet_by_name('工作表1')
for i in range (2,length+1):
    data.append([ws.cell(column=j,row=i).value for j in range (1,ws.max_column+1)])

length=156
wb = xl.load_workbook('data/left2.xlsx')
ws = wb.get_sheet_by_name('工作表1')
for i in range (2,length+1):
    data.append([ws.cell(column=j,row=i).value for j in range (1,ws.max_column+1)])


#%%
import numpy as np
l=[]
c=[]
r=[]
s=[]
t=[]
data_size=1619
for i in range (0,data_size):
    l.append(data[i][0])
    c.append(data[i][1])
    r.append(data[i][2])
    s.append(data[i][3])
    t.append(data[i][4])
L=np.array(l)[:,np.newaxis]
C=np.array(c)[:,np.newaxis]
R=np.array(r)[:,np.newaxis]
S=np.array(s)[:,np.newaxis]
T=np.array(t)[:,np.newaxis]
x=np.hstack((L,C,R))
y=np.hstack((S,T))
#%%
#preprocessing
#remove extreme data
from sklearn.preprocessing import StandardScaler
dele=[]
for i in range(0,data_size):
    if x[i][0]==0 or x[i][1]==0 or x[i][2]==0 : dele.append(i)
    elif x[i][0]>200 or x[i][1]>200 or x[i][2]>200 : dele.append(i)
    elif i>1 and x[i][1]>x[i][0]+15 or x[i][1]>x[i][2]+15:
        if x[i][1]-x[i-1][1]<0.5 and x[i][1]-x[i-1][1]>-0.5:
            y[i][0]=8
            y[i-1][0]=8
y=np.delete(y,dele,axis=0)
x=np.delete(x,dele,axis=0)
#standradize
scaler=StandardScaler()
scaler.fit(x)
print("mean:",scaler.mean_)
print("scale:",scaler.scale_)
#%% svm
from sklearn.model_selection import train_test_split
from sklearn import svm
from sklearn.metrics import accuracy_score
from sklearn.model_selection import GridSearchCV
from sklearn.preprocessing import scale

x=scale(x)
param_grid = {'C': [100,200,500,1000,5000,7500,10000],
              'gamma':[0.01,0.05,0.1,0.5,0.75,1,1.5, 2,5], }
x_train,x_test,y_train,y_test = train_test_split(x,T,test_size=0.1,random_state=0)

gclf = GridSearchCV(svm.SVC(kernel='rbf'), param_grid)
gclf.fit(x_train, y_train.ravel())
best_C=gclf.best_estimator_.C
best_gamma=gclf.best_estimator_.gamma
print("C_T:",best_C)
print("gamma_T:",best_gamma)
clf = svm.SVC(gamma=best_gamma, C=best_C)
clf.fit(x_train,y_train.ravel())
y_pdt_T = clf.predict(x_test)
acc_T=accuracy_score(y_pdt_T,y_test.ravel())
#%%
#合併
x_pdt_T = clf.predict(x)
pdt_T=np.array(x_pdt_T)[:,np.newaxis]
x_new=np.hstack((L,C,R,pdt_T))
x_train,x_test,y_train,y_test = train_test_split(x_new,T,test_size=0.1,random_state=0)

gclf = GridSearchCV(svm.SVC(kernel='rbf'), param_grid)
gclf.fit(x_train, y_train.ravel())
best_C=gclf.best_estimator_.C
best_gamma=gclf.best_estimator_.gamma
print("C_S:",best_C)
print("gamma_S:",best_gamma)
clf_S = svm.SVC(gamma=best_gamma, C=best_C)
clf_S.fit(x_train,y_train.ravel())
y_pdt_S = clf_S.predict(x_test)
acc_S=accuracy_score(y_pdt_S,y_test.ravel())
#%% porter
from sklearn_porter import Porter
porter=Porter(clf,language='C')
output=porter.export(embed_data=True)
output_file_name="svm_module.txt"
with open(output_file_name,'w') as file_obj:
    file_obj.write(output)